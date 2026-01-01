"""
Excel-JSON 양방향 동기화 스크립트

기능:
- Export: Excel 데이터를 results_verified.json 형식으로 내보내기
- Import: results_verified.json을 읽어 Excel에 새 모델 컬럼 추가

사용법:
    python sync_data.py export --sheet 국어-공통 --model "GPT-5.1"
    python sync_data.py export --sheet 국어-공통 --all-models
    python sync_data.py export --all-sheets                      # 모든 과목을 하나의 JSON 배열로 출력
    python sync_data.py export --all-sheets --output results.json
    python sync_data.py import --json problems/국어/공통/results_verified.json
    python sync_data.py import --all
    python sync_data.py list
    python sync_data.py validate
"""

import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
from copy import deepcopy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color


class PathMapper:
    """Excel 시트명과 JSON 경로 간 매핑"""

    # 시트명 -> JSON 경로 매핑
    SHEET_TO_JSON = {
        '국어-공통': 'problems/국어/공통',
        '국어-화작': 'problems/국어/화작',
        '국어-언매': 'problems/국어/언매',
        '수학-공통': 'problems/수학/공통',
        '수학-확통': 'problems/수학/확통',
        '수학-미적': 'problems/수학/미적',
        '수학-기하': 'problems/수학/기하',
        '영어': 'problems/영어',
        '한국사': 'problems/한국사',
        '물리1': 'problems/탐구/물리1',
        '화학1': 'problems/탐구/화학1',
        '생명1': 'problems/탐구/생명1',
        '사회문화': 'problems/탐구/사회문화',
    }

    def __init__(self, base_dir: Path = None):
        self.base_dir = base_dir or Path('.')
        # 역방향 매핑 생성
        self._json_to_sheet = {v: k for k, v in self.SHEET_TO_JSON.items()}

    def sheet_to_json_path(self, sheet_name: str) -> Optional[Path]:
        """시트 이름으로 JSON 폴더 경로 반환"""
        if sheet_name in self.SHEET_TO_JSON:
            return self.base_dir / self.SHEET_TO_JSON[sheet_name]
        return None

    def json_to_sheet_name(self, json_path: Path) -> Optional[str]:
        """JSON 경로로 시트 이름 반환"""
        # 경로 정규화
        rel_path = str(json_path).replace('\\', '/')
        # results_verified.json 제거
        if rel_path.endswith('/results_verified.json'):
            rel_path = rel_path[:-len('/results_verified.json')]
        elif rel_path.endswith('results_verified.json'):
            rel_path = str(Path(rel_path).parent).replace('\\', '/')

        # 매핑 검색
        for json_rel, sheet in self._json_to_sheet.items():
            if rel_path.endswith(json_rel) or json_rel in rel_path:
                return sheet
        return None

    def get_all_sheets(self) -> List[str]:
        """모든 시트 이름 반환"""
        return list(self.SHEET_TO_JSON.keys())

    def get_subject_section(self, sheet_name: str) -> Tuple[str, str]:
        """시트 이름에서 과목, 섹션 추출"""
        if '-' in sheet_name:
            parts = sheet_name.split('-', 1)
            return parts[0].strip(), parts[1].strip()
        return sheet_name.strip(), sheet_name.strip()


class ModelNameMapper:
    """모델 이름 매핑 (JSON 이름 <-> Excel 이름)"""

    def __init__(self, mapping_file: Path = None):
        self.mapping_file = mapping_file or Path('model_mapping.json')
        self.mapping = {}  # JSON name -> Excel name
        self._load_mapping()

    def _load_mapping(self):
        """매핑 파일 로드"""
        if self.mapping_file.exists():
            with open(self.mapping_file, 'r', encoding='utf-8') as f:
                self.mapping = json.load(f)

    def save_mapping(self):
        """매핑 파일 저장"""
        with open(self.mapping_file, 'w', encoding='utf-8') as f:
            json.dump(self.mapping, f, ensure_ascii=False, indent=2)

    def json_to_excel(self, json_name: str) -> str:
        """JSON 모델 이름을 Excel 컬럼 이름으로 변환"""
        return self.mapping.get(json_name, json_name)

    def excel_to_json(self, excel_name: str) -> str:
        """Excel 컬럼 이름을 JSON 모델 이름으로 변환"""
        # 역방향 검색
        for json_name, mapped_excel in self.mapping.items():
            if mapped_excel == excel_name:
                return json_name
        return excel_name

    def add_mapping(self, json_name: str, excel_name: str):
        """새 매핑 추가"""
        self.mapping[json_name] = excel_name


class ExcelHandler:
    """Excel 파일 읽기/쓰기"""

    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self._header_row_cache = {}

    def _load_workbook(self):
        """워크북 로드 (lazy loading)"""
        if self.workbook is None:
            self.workbook = load_workbook(self.excel_path)

    def get_sheet_names(self) -> List[str]:
        """모든 시트 이름 반환"""
        self._load_workbook()
        return self.workbook.sheetnames

    def _find_header_row(self, sheet_name: str) -> int:
        """헤더 행 번호 찾기 (1-based)"""
        if sheet_name in self._header_row_cache:
            return self._header_row_cache[sheet_name]

        self._load_workbook()
        ws = self.workbook[sheet_name]

        for row_idx in range(1, min(6, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and '문항 번호' in str(cell_value):
                self._header_row_cache[sheet_name] = row_idx
                return row_idx

        raise ValueError(f"'{sheet_name}' 시트에서 헤더 행을 찾을 수 없습니다.")

    def _find_score_row(self, sheet_name: str) -> int:
        """총점 행 번호 찾기 (1-based)"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)

        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and str(cell_value).strip() in ['총점', '총합', '점수']:
                return row_idx

        raise ValueError(f"'{sheet_name}' 시트에서 총점 행을 찾을 수 없습니다.")

    def get_model_columns(self, sheet_name: str) -> Dict[str, int]:
        """모델 컬럼 이름과 열 번호 반환 (1-based)"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)

        models = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                col_str = str(cell_value).strip()
                # 불필요한 컬럼 제외
                if col_str in ['문항 번호', '정답', 'nan', '']:
                    continue
                if 'Unnamed' in col_str:
                    continue
                models[col_str] = col_idx

        return models

    def get_model_answers(self, sheet_name: str, model_name: str) -> Dict[int, any]:
        """특정 모델의 문항별 답 추출"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)

        model_columns = self.get_model_columns(sheet_name)
        if model_name not in model_columns:
            raise ValueError(f"'{model_name}' 모델을 '{sheet_name}' 시트에서 찾을 수 없습니다.")

        col_idx = model_columns[model_name]
        answers = {}

        for row_idx in range(header_row + 1, score_row):
            q_num = ws.cell(row=row_idx, column=1).value
            answer = ws.cell(row=row_idx, column=col_idx).value

            if q_num is not None:
                try:
                    q_num = int(q_num)
                    answers[q_num] = answer
                except (ValueError, TypeError):
                    pass

        return answers

    def get_model_score(self, sheet_name: str, model_name: str) -> Optional[int]:
        """특정 모델의 총점 반환"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        score_row = self._find_score_row(sheet_name)

        model_columns = self.get_model_columns(sheet_name)
        if model_name not in model_columns:
            return None

        col_idx = model_columns[model_name]
        score = ws.cell(row=score_row, column=col_idx).value

        try:
            return int(score)
        except (ValueError, TypeError):
            return None

    def get_max_score(self, sheet_name: str) -> int:
        """만점 반환 (정답 열의 총점)"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)

        # 정답 열 찾기
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value and str(cell_value).strip() == '정답':
                max_score = ws.cell(row=score_row, column=col_idx).value
                try:
                    return int(max_score)
                except:
                    pass

        return 100  # 기본값

    def _find_answer_column(self, sheet_name: str) -> int:
        """정답 열 번호 찾기 (1-based)"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value and str(cell_value).strip() == '정답':
                return col_idx

        return 2  # 기본값

    def _get_correct_answers(self, sheet_name: str) -> Dict[int, any]:
        """정답 데이터 추출 {문항번호: 정답}"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)
        answer_col = self._find_answer_column(sheet_name)

        correct_answers = {}
        for row_idx in range(header_row + 1, score_row):
            q_num = ws.cell(row=row_idx, column=1).value
            correct = ws.cell(row=row_idx, column=answer_col).value
            if q_num is not None:
                try:
                    q_num = int(q_num)
                    correct_answers[q_num] = correct
                except (ValueError, TypeError):
                    pass

        return correct_answers

    def add_model_column(self, sheet_name: str, model_name: str,
                         answers: Dict[int, any], score: int,
                         position: Optional[int] = None,
                         after_model: Optional[str] = None) -> int:
        """
        새 모델 컬럼 추가

        Args:
            sheet_name: 시트 이름
            model_name: 모델 이름
            answers: {문항번호: 답} 딕셔너리
            score: 총점
            position: 삽입할 열 번호 (1-based, None이면 마지막)
            after_model: 이 모델 다음에 삽입

        Returns:
            삽입된 열 번호
        """
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)

        model_columns = self.get_model_columns(sheet_name)

        # 이미 존재하는 모델인지 확인
        if model_name in model_columns:
            raise ValueError(f"'{model_name}' 모델이 이미 존재합니다. --update 옵션을 사용하세요.")

        # 삽입 위치 결정
        if after_model and after_model in model_columns:
            insert_col = model_columns[after_model] + 1
        elif position is not None:
            insert_col = position
        else:
            # 마지막 모델 컬럼 다음
            if model_columns:
                insert_col = max(model_columns.values()) + 1
            else:
                insert_col = 3  # 정답 다음

        # 열 삽입
        ws.insert_cols(insert_col)

        # 스타일 정의
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        red_font = Font(color='FF0000')
        red_bold_font = Font(color='FF0000', bold=True)

        # 정답 데이터 가져오기
        correct_answers = self._get_correct_answers(sheet_name)

        # 헤더 설정 (볼드 + 중앙정렬)
        header_cell = ws.cell(row=header_row, column=insert_col, value=model_name)
        header_cell.font = bold_font
        header_cell.alignment = center_align

        # 답안 입력
        for row_idx in range(header_row + 1, score_row):
            q_num = ws.cell(row=row_idx, column=1).value
            if q_num is not None:
                try:
                    q_num = int(q_num)
                    cell = ws.cell(row=row_idx, column=insert_col)
                    cell.alignment = center_align

                    if q_num in answers:
                        answer = answers[q_num]
                        correct = correct_answers.get(q_num)

                        # 무응답 (-1, None, 빈 문자열)
                        if answer is None or answer == -1 or answer == "":
                            cell.value = "(포기)"
                            cell.font = red_font
                        else:
                            cell.value = answer
                            # 오답 확인
                            try:
                                if int(answer) != int(correct):
                                    cell.font = red_font
                            except (ValueError, TypeError):
                                pass
                    else:
                        # answers에 없는 문항도 포기 처리
                        cell.value = "(포기)"
                        cell.font = red_font

                except (ValueError, TypeError):
                    pass

        # 총점 입력 (중앙정렬)
        score_cell = ws.cell(row=score_row, column=insert_col, value=score)
        score_cell.alignment = center_align

        # 캐시 무효화
        self._header_row_cache.pop(sheet_name, None)

        return insert_col

    def update_model_column(self, sheet_name: str, model_name: str,
                            answers: Dict[int, any], score: int):
        """기존 모델 컬럼 업데이트"""
        self._load_workbook()
        ws = self.workbook[sheet_name]
        header_row = self._find_header_row(sheet_name)
        score_row = self._find_score_row(sheet_name)

        model_columns = self.get_model_columns(sheet_name)
        if model_name not in model_columns:
            raise ValueError(f"'{model_name}' 모델을 찾을 수 없습니다.")

        col_idx = model_columns[model_name]

        # 스타일 정의
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        red_font = Font(color='FF0000')

        # 정답 데이터 가져오기
        correct_answers = self._get_correct_answers(sheet_name)

        # 헤더 스타일 (볼드 + 중앙정렬)
        header_cell = ws.cell(row=header_row, column=col_idx)
        header_cell.font = bold_font
        header_cell.alignment = center_align

        # 답안 업데이트
        for row_idx in range(header_row + 1, score_row):
            q_num = ws.cell(row=row_idx, column=1).value
            if q_num is not None:
                try:
                    q_num = int(q_num)
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_align
                    cell.font = Font()  # 기본 폰트로 초기화

                    if q_num in answers:
                        answer = answers[q_num]
                        correct = correct_answers.get(q_num)

                        # 무응답 (-1, None, 빈 문자열)
                        if answer is None or answer == -1 or answer == "":
                            cell.value = "(포기)"
                            cell.font = red_font
                        else:
                            cell.value = answer
                            # 오답 확인
                            try:
                                if int(answer) != int(correct):
                                    cell.font = red_font
                            except (ValueError, TypeError):
                                pass
                    else:
                        cell.value = "(포기)"
                        cell.font = red_font

                except (ValueError, TypeError):
                    pass

        # 총점 업데이트 (중앙정렬)
        score_cell = ws.cell(row=score_row, column=col_idx, value=score)
        score_cell.alignment = center_align

    def save(self):
        """변경사항 저장"""
        if self.workbook:
            self.workbook.save(self.excel_path)
            print(f"저장 완료: {self.excel_path}")


class DataConverter:
    """데이터 변환 로직"""

    def __init__(self, path_mapper: PathMapper, model_mapper: ModelNameMapper):
        self.path_mapper = path_mapper
        self.model_mapper = model_mapper

    def load_questions(self, sheet_name: str) -> Dict:
        """questions.json 로드"""
        json_path = self.path_mapper.sheet_to_json_path(sheet_name)
        if not json_path:
            raise ValueError(f"'{sheet_name}' 시트에 대한 경로 매핑을 찾을 수 없습니다.")

        questions_file = json_path / 'questions.json'
        if not questions_file.exists():
            raise FileNotFoundError(f"questions.json을 찾을 수 없습니다: {questions_file}")

        with open(questions_file, 'r', encoding='utf-8') as f:
            return json.load(f)

    def excel_to_json(self, sheet_name: str, model_name: str,
                      answers: Dict[int, any], excel_handler: ExcelHandler) -> Dict:
        """
        Excel의 한 모델 컬럼 데이터를 results_verified.json 형식으로 변환

        Args:
            sheet_name: 시트 이름
            model_name: Excel의 모델 이름
            answers: {문항번호: 답}
            excel_handler: ExcelHandler 인스턴스

        Returns:
            results_verified.json 형식의 딕셔너리
        """
        questions_data = self.load_questions(sheet_name)
        subject, section = self.path_mapper.get_subject_section(sheet_name)
        json_model_name = self.model_mapper.excel_to_json(model_name)

        results = []
        total_score = 0
        correct_count = 0

        for q in questions_data['questions']:
            q_num = q['number']
            correct_answer = q['correct_answer']
            points = q['points']

            extracted = answers.get(q_num)

            # 답안 정규화
            if extracted is None or extracted == '' or str(extracted).strip() == '':
                extracted_normalized = -1
            elif str(extracted).strip() == '(포기)':
                extracted_normalized = -1
            else:
                try:
                    extracted_normalized = int(extracted)
                except (ValueError, TypeError):
                    extracted_normalized = -1

            is_correct = (extracted_normalized == correct_answer)
            if is_correct:
                total_score += points
                correct_count += 1

            results.append({
                'question_number': q_num,
                'model_name': json_model_name,
                'extracted_answer': extracted_normalized,
                'correct_answer': correct_answer,
                'is_correct': is_correct,
                'points': points,
                'needs_manual_review': False,
                'raw_response': ''
            })

        total_points = sum(q['points'] for q in questions_data['questions'])

        return {
            'subject': questions_data.get('subject', subject),
            'section': questions_data.get('section', section),
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_points': total_points,
            'total_verified': len(results),
            'correct_count': correct_count,
            'manual_review_count': 0,
            'model_scores': {json_model_name: total_score},
            'results': results
        }

    def json_to_excel(self, json_data: Dict) -> Tuple[str, Dict[int, any], int]:
        """
        results_verified.json 데이터를 Excel 형식으로 변환

        Args:
            json_data: results_verified.json 데이터

        Returns:
            (model_name, {문항번호: 답}, 총점)
        """
        # 모델 이름 추출 (첫 번째 모델)
        model_scores = json_data.get('model_scores', {})
        if not model_scores:
            raise ValueError("model_scores가 비어있습니다.")

        json_model_name = list(model_scores.keys())[0]
        excel_model_name = self.model_mapper.json_to_excel(json_model_name)
        score = model_scores[json_model_name]

        # 답안 추출
        answers = {}
        for result in json_data.get('results', []):
            if result['model_name'] == json_model_name:
                q_num = result['question_number']
                extracted = result['extracted_answer']
                # -1은 빈 셀로 변환
                if extracted == -1:
                    answers[q_num] = None
                else:
                    answers[q_num] = extracted

        return excel_model_name, answers, score


class SyncManager:
    """동기화 관리"""

    def __init__(self, excel_path: Path, problems_dir: Path = None,
                 model_mapping_path: Path = None):
        self.excel_path = Path(excel_path)
        self.problems_dir = problems_dir or Path('problems')

        self.path_mapper = PathMapper(base_dir=Path('.'))
        self.model_mapper = ModelNameMapper(model_mapping_path)
        self.excel_handler = ExcelHandler(self.excel_path)
        self.converter = DataConverter(self.path_mapper, self.model_mapper)

    def export_to_json(self, sheet_name: str, model_name: str,
                       output_path: Path = None) -> Path:
        """
        Excel -> JSON 내보내기

        Args:
            sheet_name: 시트 이름
            model_name: 모델 이름
            output_path: 출력 경로 (None이면 기본 경로)

        Returns:
            저장된 파일 경로
        """
        # 답안 추출
        answers = self.excel_handler.get_model_answers(sheet_name, model_name)

        # JSON 변환
        json_data = self.converter.excel_to_json(
            sheet_name, model_name, answers, self.excel_handler
        )

        # 출력 경로 결정
        if output_path is None:
            json_dir = self.path_mapper.sheet_to_json_path(sheet_name)
            if not json_dir:
                raise ValueError(f"'{sheet_name}'에 대한 경로 매핑이 없습니다.")
            output_path = json_dir / 'results_verified.json'

        # 기존 파일이 있으면 병합
        if output_path.exists():
            with open(output_path, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)

            # 기존 모델 점수와 결과 병합
            json_model_name = self.model_mapper.excel_to_json(model_name)
            existing_data['model_scores'][json_model_name] = json_data['model_scores'][json_model_name]

            # 기존 results에서 같은 모델 결과 제거 후 추가
            existing_data['results'] = [
                r for r in existing_data['results']
                if r['model_name'] != json_model_name
            ]
            existing_data['results'].extend(json_data['results'])
            existing_data['timestamp'] = json_data['timestamp']

            json_data = existing_data

        # 저장
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)

        print(f"내보내기 완료: {output_path}")
        return output_path

    def import_from_json(self, json_path: Path,
                         position: Optional[int] = None,
                         after_model: Optional[str] = None,
                         update_existing: bool = False,
                         excel_name: Optional[str] = None) -> bool:
        """
        JSON -> Excel 가져오기

        Args:
            json_path: results_verified.json 경로
            position: 삽입할 열 위치
            after_model: 이 모델 다음에 삽입
            update_existing: 기존 데이터 업데이트 여부
            excel_name: Excel에서 사용할 모델 이름 (None이면 매핑 사용)

        Returns:
            성공 여부
        """
        json_path = Path(json_path)
        if not json_path.exists():
            print(f"파일을 찾을 수 없습니다: {json_path}")
            return False

        # JSON 로드
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)

        # 시트 이름 결정
        sheet_name = self.path_mapper.json_to_sheet_name(json_path)
        if not sheet_name:
            print(f"경고: '{json_path}'에 대한 시트 매핑을 찾을 수 없습니다.")
            # subject/section으로 시트 이름 추론
            subject = json_data.get('subject', '')
            section = json_data.get('section', '')
            if subject and section and subject != section:
                sheet_name = f"{subject}-{section}"
            elif subject:
                sheet_name = subject
            else:
                print("시트 이름을 결정할 수 없습니다.")
                return False

        # 시트 존재 확인
        if sheet_name not in self.excel_handler.get_sheet_names():
            print(f"시트를 찾을 수 없습니다: {sheet_name}")
            return False

        # 데이터 변환
        json_model_name, answers, score = self.converter.json_to_excel(json_data)

        # Excel 모델 이름 결정
        if excel_name:
            model_name = excel_name
        else:
            model_name = self.model_mapper.json_to_excel(json_model_name)

        # 기존 모델 확인
        existing_models = self.excel_handler.get_model_columns(sheet_name)

        if model_name in existing_models:
            if update_existing:
                self.excel_handler.update_model_column(sheet_name, model_name, answers, score)
                print(f"업데이트 완료: {sheet_name} / {model_name} ({score}점)")
            else:
                print(f"'{model_name}' 모델이 이미 존재합니다. --update 옵션을 사용하세요.")
                return False
        else:
            self.excel_handler.add_model_column(
                sheet_name, model_name, answers, score,
                position=position, after_model=after_model
            )
            print(f"추가 완료: {sheet_name} / {model_name} ({score}점)")

        return True

    def import_all(self, update_existing: bool = False) -> int:
        """모든 results_verified.json 가져오기"""
        count = 0
        for sheet_name in self.path_mapper.get_all_sheets():
            json_dir = self.path_mapper.sheet_to_json_path(sheet_name)
            if json_dir:
                json_file = json_dir / 'results_verified.json'
                if json_file.exists():
                    if self.import_from_json(json_file, update_existing=update_existing):
                        count += 1

        if count > 0:
            self.excel_handler.save()

        return count

    def export_all_sheets_to_json(self, output_path: Path = None,
                                     model_name: str = None,
                                     all_models: bool = False) -> Path:
        """
        모든 시트를 하나의 JSON 파일로 내보내기 (객체 배열 형태)

        Args:
            output_path: 출력 경로 (None이면 all_results.json)
            model_name: 특정 모델만 내보내기
            all_models: 모든 모델 내보내기

        Returns:
            저장된 파일 경로
        """
        all_data = []

        for sheet_name in self.path_mapper.get_all_sheets():
            # 시트 존재 확인
            if sheet_name not in self.excel_handler.get_sheet_names():
                continue

            # questions.json 존재 확인
            json_dir = self.path_mapper.sheet_to_json_path(sheet_name)
            if not json_dir or not (json_dir / 'questions.json').exists():
                continue

            try:
                # 모델 목록 결정
                if model_name:
                    models_to_export = [model_name]
                elif all_models:
                    models_to_export = list(self.excel_handler.get_model_columns(sheet_name).keys())
                else:
                    models_to_export = list(self.excel_handler.get_model_columns(sheet_name).keys())

                for model in models_to_export:
                    try:
                        answers = self.excel_handler.get_model_answers(sheet_name, model)
                        json_data = self.converter.excel_to_json(
                            sheet_name, model, answers, self.excel_handler
                        )
                        # 깔끔한 형태로 재구성
                        json_model_name = list(json_data['model_scores'].keys())[0]
                        # results에서 불필요한 필드 제거
                        clean_results = [
                            {
                                'question_number': r['question_number'],
                                'extracted_answer': r['extracted_answer'],
                                'correct_answer': r['correct_answer'],
                                'is_correct': r['is_correct'],
                                'points': r['points'],
                            }
                            for r in json_data['results']
                        ]
                        clean_data = {
                            'sheet_name': sheet_name,
                            'subject': json_data['subject'],
                            'section': json_data['section'],
                            'model_name': json_model_name,
                            'score': json_data['model_scores'][json_model_name],
                            'total_points': json_data['total_points'],
                            'correct_count': json_data['correct_count'],
                            'total_questions': json_data['total_verified'],
                            'results': clean_results,
                        }
                        all_data.append(clean_data)
                    except Exception as e:
                        print(f"경고: {sheet_name}/{model} 내보내기 실패 - {e}")

            except Exception as e:
                print(f"경고: {sheet_name} 처리 실패 - {e}")

        # 출력 경로 결정
        if output_path is None:
            output_path = Path('all_results.json')

        # 저장
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(all_data, f, ensure_ascii=False, indent=2)

        print(f"내보내기 완료: {output_path} ({len(all_data)}개 항목)")
        return output_path

    def list_models(self, sheet_name: str = None) -> Dict[str, List[str]]:
        """모델 목록 반환"""
        result = {}

        if sheet_name:
            sheets = [sheet_name]
        else:
            sheets = self.excel_handler.get_sheet_names()

        for sheet in sheets:
            if sheet in self.path_mapper.SHEET_TO_JSON or sheet_name:
                try:
                    models = list(self.excel_handler.get_model_columns(sheet).keys())
                    result[sheet] = models
                except:
                    pass

        return result

    def validate(self, sheet_name: str = None) -> List[str]:
        """데이터 일관성 검증"""
        issues = []

        if sheet_name:
            sheets = [sheet_name]
        else:
            sheets = self.path_mapper.get_all_sheets()

        for sheet in sheets:
            json_dir = self.path_mapper.sheet_to_json_path(sheet)
            if not json_dir:
                continue

            # questions.json 확인
            questions_file = json_dir / 'questions.json'
            if not questions_file.exists():
                issues.append(f"[{sheet}] questions.json 없음: {questions_file}")
                continue

            with open(questions_file, 'r', encoding='utf-8') as f:
                questions = json.load(f)

            expected_count = len(questions['questions'])
            expected_total = sum(q['points'] for q in questions['questions'])

            # Excel 데이터 확인
            try:
                model_columns = self.excel_handler.get_model_columns(sheet)
                max_score = self.excel_handler.get_max_score(sheet)

                if max_score != expected_total:
                    issues.append(f"[{sheet}] 만점 불일치: Excel={max_score}, JSON={expected_total}")

                for model_name in model_columns:
                    answers = self.excel_handler.get_model_answers(sheet, model_name)
                    if len(answers) != expected_count:
                        issues.append(
                            f"[{sheet}] {model_name}: 문항 수 불일치 "
                            f"(Excel={len(answers)}, JSON={expected_count})"
                        )
            except Exception as e:
                issues.append(f"[{sheet}] 검증 오류: {e}")

        return issues


def main():
    parser = argparse.ArgumentParser(
        description='Excel-JSON 양방향 동기화 도구',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  # JSON -> Excel 가져오기
  python sync_data.py import --json problems/국어/공통/results_verified.json
  python sync_data.py import --all

  # Excel -> JSON 내보내기 (단일 시트)
  python sync_data.py export --sheet 국어-공통 --model "GPT-5.1"
  python sync_data.py export --sheet 국어-공통 --all-models

  # Excel -> JSON 내보내기 (모든 시트를 하나의 JSON 배열로)
  python sync_data.py export --all-sheets
  python sync_data.py export --all-sheets --output all_results.json
  python sync_data.py export --all-sheets --model "GPT-5.1"  # 특정 모델만

  # 모델 목록 확인
  python sync_data.py list
  python sync_data.py list --sheet 국어-공통

  # 검증
  python sync_data.py validate
        """
    )

    subparsers = parser.add_subparsers(dest='command', help='명령')

    # Export 명령
    export_parser = subparsers.add_parser('export', help='Excel -> JSON 내보내기')
    export_parser.add_argument('--sheet', help='시트 이름 (예: 국어-공통)')
    export_parser.add_argument('--all-sheets', action='store_true', help='모든 시트를 하나의 JSON 배열로 내보내기')
    export_parser.add_argument('--model', help='모델 이름')
    export_parser.add_argument('--all-models', action='store_true', help='모든 모델 내보내기')
    export_parser.add_argument('--output', help='출력 파일 경로')

    # Import 명령
    import_parser = subparsers.add_parser('import', help='JSON -> Excel 가져오기')
    import_parser.add_argument('--json', help='results_verified.json 경로')
    import_parser.add_argument('--all', action='store_true', help='모든 JSON 파일 가져오기')
    import_parser.add_argument('--position', type=int, help='삽입할 열 위치')
    import_parser.add_argument('--after', help='이 모델 다음에 삽입')
    import_parser.add_argument('--update', action='store_true', help='기존 데이터 업데이트')
    import_parser.add_argument('--excel-name', help='Excel에서 사용할 모델 이름')

    # List 명령
    list_parser = subparsers.add_parser('list', help='모델 목록 확인')
    list_parser.add_argument('--sheet', help='특정 시트만')

    # Validate 명령
    validate_parser = subparsers.add_parser('validate', help='데이터 검증')
    validate_parser.add_argument('--sheet', help='특정 시트만')

    # 공통 옵션
    parser.add_argument('--excel', default='2026 수능 LLM 풀이.xlsx',
                        help='Excel 파일 경로')
    parser.add_argument('--mapping', default='model_mapping.json',
                        help='모델 이름 매핑 파일')

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return

    # SyncManager 초기화
    sync = SyncManager(
        excel_path=Path(args.excel),
        model_mapping_path=Path(args.mapping)
    )

    # 명령 실행
    if args.command == 'export':
        if args.all_sheets:
            # 모든 시트를 하나의 JSON 배열로 내보내기
            output = Path(args.output) if args.output else None
            sync.export_all_sheets_to_json(
                output_path=output,
                model_name=args.model,
                all_models=args.all_models or (args.model is None)
            )
        elif args.sheet:
            if args.all_models:
                models = sync.list_models(args.sheet).get(args.sheet, [])
                for model in models:
                    sync.export_to_json(args.sheet, model)
            elif args.model:
                output = Path(args.output) if args.output else None
                sync.export_to_json(args.sheet, args.model, output)
            else:
                print("--model 또는 --all-models 옵션을 지정하세요.")
        else:
            print("--sheet 또는 --all-sheets 옵션을 지정하세요.")

    elif args.command == 'import':
        if args.all:
            count = sync.import_all(update_existing=args.update)
            print(f"\n총 {count}개 파일 가져오기 완료")
        elif args.json:
            success = sync.import_from_json(
                Path(args.json),
                position=args.position,
                after_model=args.after,
                update_existing=args.update,
                excel_name=args.excel_name
            )
            if success:
                sync.excel_handler.save()
        else:
            print("--json 또는 --all 옵션을 지정하세요.")

    elif args.command == 'list':
        models = sync.list_models(args.sheet)
        for sheet, model_list in models.items():
            print(f"\n[{sheet}] ({len(model_list)}개 모델)")
            for model in model_list:
                score = sync.excel_handler.get_model_score(sheet, model)
                print(f"  - {model}: {score}점")

    elif args.command == 'validate':
        issues = sync.validate(args.sheet)
        if issues:
            print("검증 결과: 문제 발견")
            for issue in issues:
                print(f"  - {issue}")
        else:
            print("검증 결과: 정상")


if __name__ == '__main__':
    main()
